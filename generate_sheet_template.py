import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import os

def build_excel_template(file_path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    DATA_FONT = Font(name="Calibri", size=11)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
    THIN_BORDER = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # 1. Sheet: registered
    ws_registered = wb.create_sheet(title="registered")
    reg_headers = ["Timestamp", "Registration ID / QR Code", "ชื่อ-นามสกุล", "อีเมล", "เบอร์โทร", "บริษัท/หน่วยงาน/องค์กร", "แผนก", "ประเภทพนักงาน", "Ticket Number", "Lucky Draw Number", "สถานะ"]
    ws_registered.append(reg_headers)
    
    reg_sample_data = [
        ["2026-07-23 10:00:00", "EVT-100001", "สมชาย ใจดี", "somchai@example.com", "0812345678", "บริษัท เทคโนโลยี จำกัด", "ไอที", "ประจำ", "-", "-", "รอการอนุมัติ"],
        ["2026-07-23 10:05:00", "EVT-100002", "สมหญิง รักดี", "somying@example.com", "0823456789", "องค์กรภาครัฐ", "การเงิน", "สัญญาจ้าง", "-", "-", "รอการอนุมัติ"],
        ["2026-07-23 10:10:00", "EVT-100003", "กิตติพงษ์ มั่นคง", "kittipong@example.com", "0834567890", "หน่วยงานอิสระ", "การตลาด", "ชั่วคราว", "-", "-", "รอการอนุมัติ"],
    ]
    for row in reg_sample_data:
        ws_registered.append(row)

    # 2. Sheet: apploval
    ws_apploval = wb.create_sheet(title="apploval")
    app_headers = ["Timestamp", "Registration ID / QR Code", "ชื่อ-นามสกุล", "อีเมล", "เบอร์โทร", "บริษัท/หน่วยงาน/องค์กร", "แผนก", "ประเภทพนักงาน", "Ticket Number", "Lucky Draw Number", "สถานะ"]
    ws_apploval.append(app_headers)
    
    app_sample_data = [
        ["2026-07-23 09:30:00", "EVT-100000", "อนันต์ ขยันยิ่ง", "anan@example.com", "0898765432", "บริษัท เทคโนโลยี จำกัด", "ไอที", "ประจำ", "A00001", "A00001", "อนุมัติ"],
        ["2026-07-23 09:35:00", "EVT-100004", "วิภาดา สุขใจ", "wiphada@example.com", "0887654321", "องค์กรภาครัฐ", "การเงิน", "ประจำ", "B00001", "B00001", "อนุมัติ"],
    ]
    for row in app_sample_data:
        ws_apploval.append(row)

    # 3. Sheet: club
    ws_club = wb.create_sheet(title="club")
    club_headers = ["ชื่อบริษัท/หน่วยงาน/องค์กร", "สถานะใช้งาน"]
    ws_club.append(club_headers)
    
    club_sample_data = [
        ["บริษัท เทคโนโลยี จำกัด", "ใช้งาน"],
        ["องค์กรภาครัฐ", "ใช้งาน"],
        ["หน่วยงานอิสระ", "ใช้งาน"],
        ["สมาคมธุรกิจดิจิทัล", "ใช้งาน"],
        ["สถาบันการศึกษา", "ใช้งาน"],
    ]
    for row in club_sample_data:
        ws_club.append(row)

    # 4. Sheet: winer
    ws_winer = wb.create_sheet(title="winer")
    win_headers = ["Lucky Draw Number", "Ticket Number", "ชื่อ-นามสกุล", "บริษัท/หน่วยงาน/องค์กร", "ชื่อรางวัลที่ได้รับ", "เวลาที่ได้รับรางวัล"]
    ws_winer.append(win_headers)
    
    win_sample_data = [
        ["A00001", "A00001", "อนันต์ ขยันยิ่ง", "บริษัท เทคโนโลยี จำกัด", "รางวัลใหญ่ iPad Pro", "2026-07-23 15:00:00"],
    ]
    for row in win_sample_data:
        ws_winer.append(row)

    # Apply Styling to all sheets
    for ws in wb.worksheets:
        # Header Row Styling
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        
        # Data Rows Styling & Auto Column Width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = DATA_FONT
                cell.alignment = LEFT_ALIGN
                cell.border = THIN_BORDER

        # Auto Column Widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(file_path)
    print(f"Successfully generated Excel template at: {file_path}")

    # Generate CSV files in gas/ directory as well
    gas_dir = os.path.join(os.path.dirname(os.path.abspath(file_path)), "gas")
    sheets_csv = {
        "registered": (reg_headers, reg_sample_data),
        "apploval": (app_headers, app_sample_data),
        "club": (club_headers, club_sample_data),
        "winer": (win_headers, win_sample_data),
    }
    
    for sheet_name, (headers, data) in sheets_csv.items():
        csv_path = os.path.join(gas_dir, f"{sheet_name}.csv")
        with open(csv_path, mode="w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)
        print(f"Generated CSV: {csv_path}")

if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    build_excel_template(os.path.join(current_dir, "Event_Registration_Google_Sheet_Template.xlsx"))
