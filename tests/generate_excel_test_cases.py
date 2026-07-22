import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE_PATH = r"d:\Project\event-platform\tests\test_cases.xlsx"

def create_excel_test_cases():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Automation Test Cases"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep Navy
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        "TestCaseID",
        "Suite",
        "TestCaseName",
        "Method",
        "Endpoint",
        "PayloadJSON",
        "RequireAuth",
        "ExpectedStatus",
        "ExpectedField"
    ]

    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    test_cases = [
        # Suite 1: Auth
        ["TC-001", "Auth", "Admin Valid Login", "POST", "/api/auth/login", '{"username": "admin", "password": "password123"}', "FALSE", 200, "data.token"],
        ["TC-002", "Auth", "Reject Invalid Password Login", "POST", "/api/auth/login", '{"username": "admin", "password": "wrongpassword"}', "FALSE", 401, "message"],
        
        # Suite 2: Events CRUD
        ["TC-003", "Events", "Create New Event", "POST", "/api/events", '{"name": "Tech Gala Expo ${timestamp}", "description": "Annual Technology Innovation Summit", "venue": "Grand Hall 1", "startDate": "2026-12-01T09:00:00.000Z", "endDate": "2026-12-01T18:00:00.000Z", "status": "ACTIVE"}', "TRUE", 201, "data.id"],
        ["TC-004", "Events", "Update Event Status to Active/Public", "PUT", "/api/events/${eventId}", '{"status": "ACTIVE"}', "TRUE", 200, "data.id"],
        ["TC-005", "Events", "Get Event Detail", "GET", "/api/events/${eventId}", "", "FALSE", 200, "data.name"],
        
        # Suite 3: CheckinPoints CRUD
        ["TC-006", "CheckinPoints", "Get Entrance Points List", "GET", "/api/checkin/points?eventId=${eventId}", "", "FALSE", 200, "data"],
        ["TC-007", "CheckinPoints", "Create New Entrance Gate", "POST", "/api/checkin/points", '{"eventId": "${eventId}", "name": "VIP Entrance Gate 1", "location": "Hall A North", "isActive": true, "sortOrder": 1}', "TRUE", 201, "data.id"],
        
        # Suite 4: Multi-Attendee Registration & Gate Check-in
        ["TC-008", "Registration", "Register New Attendee (Guest 1)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Somchai Jaidee (Guest 1)", "email": "somchai.${timestamp}@example.com", "phone": "0811111111", "company": "Tech Corp"}', "FALSE", 201, "data.id"],
        ["TC-009", "Checkin", "Perform Gate Check-in (Guest 1)", "POST", "/api/checkin", '{"registrationId": "${registrationId}", "checkinPointId": "${pointId}", "method": "MANUAL"}', "FALSE", 201, "data.checkin.id"],
        
        ["TC-010", "Registration", "Register New Attendee (Guest 2)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Siriporn Boonmee (Guest 2)", "email": "siriporn.${timestamp}@example.com", "phone": "0822222222", "company": "Innovate Ltd"}', "FALSE", 201, "data.id"],
        ["TC-011", "Checkin", "Perform Gate Check-in (Guest 2)", "POST", "/api/checkin", '{"registrationId": "${registrationId}", "checkinPointId": "${pointId}", "method": "MANUAL"}', "FALSE", 201, "data.checkin.id"],
        
        ["TC-012", "Registration", "Register New Attendee (Guest 3)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Wichai Chaichana (Guest 3)", "email": "wichai.${timestamp}@example.com", "phone": "0833333333", "company": "Digital Group"}', "FALSE", 201, "data.id"],
        ["TC-013", "Checkin", "Perform Gate Check-in (Guest 3)", "POST", "/api/checkin", '{"registrationId": "${registrationId}", "checkinPointId": "${pointId}", "method": "MANUAL"}', "FALSE", 201, "data.checkin.id"],
        
        ["TC-014", "Registration", "Register New Attendee (Guest 4)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Anan Srisuk (Guest 4)", "email": "anan.${timestamp}@example.com", "phone": "0844444444", "company": "Software Inc"}', "FALSE", 201, "data.id"],
        ["TC-015", "Checkin", "Perform Gate Check-in (Guest 4)", "POST", "/api/checkin", '{"registrationId": "${registrationId}", "checkinPointId": "${pointId}", "method": "MANUAL"}', "FALSE", 201, "data.checkin.id"],
        
        ["TC-016", "Registration", "Register New Attendee (Guest 5)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Kanya Wattana (Guest 5)", "email": "kanya.${timestamp}@example.com", "phone": "0855555555", "company": "Cloud Systems"}', "FALSE", 201, "data.id"],
        ["TC-017", "Registration", "Register New Attendee (Guest 6)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Nattapong Somboon (Guest 6)", "email": "nattapong.${timestamp}@example.com", "phone": "0866666666", "company": "Cyber Security"}', "FALSE", 201, "data.id"],
        ["TC-018", "Registration", "Register New Attendee (Guest 7)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Piyada Rattanaporn (Guest 7)", "email": "piyada.${timestamp}@example.com", "phone": "0877777777", "company": "Data Analytics"}', "FALSE", 201, "data.id"],
        ["TC-019", "Registration", "Register New Attendee (Guest 8)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Thanakorn Prasert (Guest 8)", "email": "thanakorn.${timestamp}@example.com", "phone": "0888888888", "company": "AI Labs"}', "FALSE", 201, "data.id"],
        ["TC-020", "Registration", "Register New Attendee (Guest 9)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Chutima Wongwian (Guest 9)", "email": "chutima.${timestamp}@example.com", "phone": "0899999991", "company": "Media Network"}', "FALSE", 201, "data.id"],
        ["TC-021", "Registration", "Register New Attendee (Guest 10)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Phopkrit Panich (Guest 10)", "email": "phopkrit.${timestamp}@example.com", "phone": "0899999992", "company": "Finance Hub"}', "FALSE", 201, "data.id"],
        
        ["TC-022", "Registration", "Register New Attendee (Guest 11)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Suthep Ruangrit (Guest 11)", "email": "suthep.${timestamp}@example.com", "phone": "0899999993", "company": "Logistics Pro"}', "FALSE", 201, "data.id"],
        ["TC-023", "Registration", "Register New Attendee (Guest 12)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Rungtiwa Kaewmanee (Guest 12)", "email": "rungtiwa.${timestamp}@example.com", "phone": "0899999994", "company": "Creative Studio"}', "FALSE", 201, "data.id"],
        ["TC-024", "Registration", "Register New Attendee (Guest 13)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Banthit Charoen (Guest 13)", "email": "banthit.${timestamp}@example.com", "phone": "0899999995", "company": "Mobile Apps"}', "FALSE", 201, "data.id"],
        ["TC-025", "Registration", "Register New Attendee (Guest 14)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Wanida Saetang (Guest 14)", "email": "wanida.${timestamp}@example.com", "phone": "0899999996", "company": "E-Commerce Group"}', "FALSE", 201, "data.id"],
        ["TC-026", "Registration", "Register New Attendee (Guest 15)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Kraisorn Decha (Guest 15)", "email": "kraisorn.${timestamp}@example.com", "phone": "0899999997", "company": "Robotics Corp"}', "FALSE", 201, "data.id"],
        ["TC-027", "Registration", "Register New Attendee (Guest 16)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Worawat Songsak (Guest 16)", "email": "worawat.${timestamp}@example.com", "phone": "0899999998", "company": "Gaming World"}', "FALSE", 201, "data.id"],
        ["TC-028", "Registration", "Register New Attendee (Guest 17)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Jintana Phusit (Guest 17)", "email": "jintana.${timestamp}@example.com", "phone": "0812345601", "company": "Health Tech"}', "FALSE", 201, "data.id"],
        ["TC-029", "Registration", "Register New Attendee (Guest 18)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Prasit Manop (Guest 18)", "email": "prasit.${timestamp}@example.com", "phone": "0812345602", "company": "Green Energy"}', "FALSE", 201, "data.id"],
        ["TC-030", "Registration", "Register New Attendee (Guest 19)", "POST", "/api/registrations", '{"eventId": "${eventId}", "fullName": "Rattana Chokdee (Guest 19)", "email": "rattana.${timestamp}@example.com", "phone": "0812345603", "company": "Smart City"}', "FALSE", 201, "data.id"],

        # Suite 5: Prize Creation & Lucky Draw Engine
        ["TC-031", "Prizes", "Create Prize for Event", "POST", "/api/prizes", '{"eventId": "${eventId}", "name": "Grand Prize iPhone 16 Pro", "description": "Top Lucky Draw Prize", "quantity": 5, "sortOrder": 1}', "TRUE", 201, "data.id"],
        ["TC-032", "LuckyDraw", "Fetch Prizes & Eligible Count", "GET", "/api/prizes?eventId=${eventId}", "", "FALSE", 200, "data"],
        ["TC-033", "LuckyDraw", "Start Lucky Draw Session", "POST", "/api/draws/start", '{"eventId": "${eventId}", "prizeId": "${prizeId}", "drawCount": 1}', "TRUE", 201, "data.id"],
        ["TC-034", "LuckyDraw", "Execute Fisher-Yates Spin Draw (1 Winner)", "POST", "/api/draws/${drawSessionId}/spin", '{"count": 1}', "FALSE", 200, "data.winners"],

        # Suite 6: Winner Management & Quota Sync
        ["TC-035", "WinnerMgmt", "Update Winner Status to ACCEPTED", "PUT", "/api/draws/winners/${winnerId}", '{"status": "ACCEPTED"}', "TRUE", 200, "data.status"],
        ["TC-036", "WinnerMgmt", "Delete Winner & Recalculate Quota", "DELETE", "/api/draws/winners/${winnerId}", "", "TRUE", 200, "message"]
    ]

    for tc in test_cases:
        ws.append(tc)

    for row in ws.iter_rows(min_row=2, max_row=len(test_cases) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = align_left
            cell.font = Font(name="Calibri", size=10)

    col_widths = {
        "A": 12, "B": 15, "C": 38, "D": 10, "E": 40, "F": 55, "G": 14, "H": 15, "I": 15
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(FILE_PATH)
    print(f"Excel Test Cases Template generated successfully at: {FILE_PATH}")

if __name__ == "__main__":
    create_excel_test_cases()
