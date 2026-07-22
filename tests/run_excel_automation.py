import os
import sys
import time
import json
import urllib.request
import urllib.error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===================================
# API Base URL Configuration
# ===================================
ENV_TARGETS = {
    "prod": "https://event-platform-api.onrender.com",
    "production": "https://event-platform-api.onrender.com",
    "test": "http://localhost:4000",
    "local": "http://localhost:4000",
    "dev": "http://localhost:4000",
}

cli_arg = (sys.argv[1] if len(sys.argv) > 1 else "").lower().strip()
if cli_arg in ENV_TARGETS:
    API_BASE = ENV_TARGETS[cli_arg]
else:
    API_BASE = os.getenv("API_BASE") or os.getenv("TEST_API_URL") or os.getenv("NEXT_PUBLIC_API_URL") or "http://localhost:4000"

INPUT_EXCEL = r"d:\Project\event-platform\tests\test_cases.xlsx"
OUTPUT_EXCEL = r"d:\Project\event-platform\tests\test_results.xlsx"

# Global runtime variables store
context_vars = {
    "adminToken": "",
    "eventId": "",
    "prizeId": "",
    "pointId": "",
    "registrationId": "",
    "drawSessionId": "",
    "winnerId": "",
}

def replace_vars(text: str) -> str:
    if not text:
        return ""
    res = str(text)
    res = res.replace("${timestamp}", str(int(time.time() * 1000)))
    for k, v in context_vars.items():
        res = res.replace(f"${{{k}}}", str(v))
    return res

def execute_http_request(method, url, payload_str, require_auth):
    headers = {"Content-Type": "application/json"}
    if require_auth and context_vars["adminToken"]:
        headers["Authorization"] = f"Bearer {context_vars['adminToken']}"

    data_bytes = payload_str.encode('utf-8') if payload_str else None
    
    req = urllib.request.Request(url, data=data_bytes, headers=headers, method=method.upper())
    
    start_time = time.time()
    try:
        with urllib.request.urlopen(req) as resp:
            elapsed_ms = int((time.time() - start_time) * 1000)
            status_code = resp.status
            body_str = resp.read().decode('utf-8')
            try:
                body_json = json.loads(body_str)
            except:
                body_json = body_str
            return status_code, body_json, elapsed_ms
    except urllib.error.HTTPError as e:
        elapsed_ms = int((time.time() - start_time) * 1000)
        body_str = e.read().decode('utf-8')
        try:
            body_json = json.loads(body_str)
        except:
            body_json = body_str
        return e.code, body_json, elapsed_ms
    except Exception as ex:
        elapsed_ms = int((time.time() - start_time) * 1000)
        return 500, {"error": str(ex)}, elapsed_ms

def extract_json_path(obj, path_str):
    if not path_str or not isinstance(obj, dict):
        return None
    parts = path_str.split('.')
    curr = obj
    for p in parts:
        if isinstance(curr, dict) and p in curr:
            curr = curr[p]
        else:
            return None
    return curr

def run_excel_automation():
    print("\n===============================================================")
    print("EXCEL DATA-DRIVEN AUTOMATION TEST RUNNER")
    print("===============================================================\n")

    if not os.path.exists(INPUT_EXCEL):
        print(f"Error: Input Excel file not found at {INPUT_EXCEL}")
        sys.exit(1)

    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws = wb.active

    # Add Result Columns
    headers = [cell.value for cell in ws[1]]
    headers.extend(["TestResult", "ActualStatus", "ResponseTimeMs", "ExecutionLog"])
    
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Test Execution Results"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    out_ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = out_ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    pass_font = Font(name="Arial", size=10, bold=True, color="166534")
    fail_font = Font(name="Arial", size=10, bold=True, color="991B1B")

    pass_count = 0
    fail_count = 0
    total_count = 0

    for row_idx in range(2, ws.max_row + 1):
        tc_id = ws.cell(row=row_idx, column=1).value
        suite = ws.cell(row=row_idx, column=2).value
        tc_name = ws.cell(row=row_idx, column=3).value
        method = ws.cell(row=row_idx, column=4).value
        raw_endpoint = ws.cell(row=row_idx, column=5).value
        raw_payload = ws.cell(row=row_idx, column=6).value
        req_auth = str(ws.cell(row=row_idx, column=7).value).upper() == "TRUE"
        exp_status = int(ws.cell(row=row_idx, column=8).value)
        exp_field = ws.cell(row=row_idx, column=9).value

        if not tc_id:
            continue

        total_count += 1
        endpoint = replace_vars(raw_endpoint)
        payload = replace_vars(raw_payload)
        full_url = f"{API_BASE}{endpoint}"

        status_code, body_json, duration_ms = execute_http_request(method, full_url, payload, req_auth)

        # Context Variable Extraction logic
        if tc_id == "TC-001" and status_code == 200 and isinstance(body_json, dict):
            context_vars["adminToken"] = extract_json_path(body_json, "data.token") or ""
        elif tc_id == "TC-003" and status_code == 200 and isinstance(body_json, dict):
            data_arr = body_json.get("data", [])
            if data_arr and isinstance(data_arr, list):
                context_vars["eventId"] = data_arr[0].get("id", "")
        elif tc_id == "TC-005" and status_code == 200 and isinstance(body_json, dict):
            data_arr = body_json.get("data", [])
            if data_arr and isinstance(data_arr, list):
                context_vars["pointId"] = data_arr[0].get("id", "")
        elif tc_id == "TC-006" and status_code in [200, 201] and isinstance(body_json, dict):
            context_vars["pointId"] = extract_json_path(body_json, "data.id") or context_vars["pointId"]
        elif tc_id == "TC-007" and status_code in [200, 201] and isinstance(body_json, dict):
            context_vars["registrationId"] = extract_json_path(body_json, "data.id") or ""
        elif tc_id == "TC-009" and status_code == 200 and isinstance(body_json, dict):
            data_arr = body_json.get("data", [])
            if data_arr and isinstance(data_arr, list):
                context_vars["prizeId"] = data_arr[0].get("id", "")
        elif tc_id == "TC-010" and status_code in [200, 201] and isinstance(body_json, dict):
            context_vars["drawSessionId"] = extract_json_path(body_json, "data.id") or ""
        elif tc_id == "TC-011" and status_code == 200:
            # Fetch winners list from session details to populate winnerId
            s_status, s_body, _ = execute_http_request("GET", f"{API_BASE}/api/draws/{context_vars['drawSessionId']}", "", False)
            if s_status == 200 and isinstance(s_body, dict):
                winners = extract_json_path(s_body, "data.winners")
                if winners and isinstance(winners, list):
                    context_vars["winnerId"] = winners[0].get("id", "")

        # Assertion
        is_passed = True
        log_msg = "OK"

        if status_code != exp_status:
            is_passed = False
            log_msg = f"Status Mismatch: Expected {exp_status}, Got {status_code}"
        elif exp_field:
            field_val = extract_json_path(body_json, exp_field)
            if field_val is None:
                is_passed = False
                log_msg = f"Expected Field '{exp_field}' missing in response"

        if is_passed:
            pass_count += 1
            print(f"  PASS: {tc_id} ({suite}): {tc_name} ({duration_ms}ms)")
        else:
            fail_count += 1
            print(f"  FAIL: {tc_id} ({suite}): {tc_name} ({duration_ms}ms) -> {log_msg}")

        row_data = [
            tc_id, suite, tc_name, method, endpoint, payload,
            "TRUE" if req_auth else "FALSE", exp_status, exp_field,
            "PASS" if is_passed else "FAIL",
            status_code,
            duration_ms,
            log_msg if not is_passed else json.dumps(body_json, ensure_ascii=False)[:200]
        ]
        
        out_ws.append(row_data)
        
        # Colorize result cell
        res_cell = out_ws.cell(row=out_ws.max_row, column=10)
        res_cell.fill = pass_fill if is_passed else fail_fill
        res_cell.font = pass_font if is_passed else fail_font

    for col in out_ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        out_ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    out_wb.save(OUTPUT_EXCEL)

    print("\n===============================================================")
    print(f"SUMMARY: {pass_count}/{total_count} PASSED ({((pass_count/total_count)*100):.1f}%)")
    print(f"Excel Execution Report saved at: {OUTPUT_EXCEL}")
    print("===============================================================\n")

if __name__ == "__main__":
    run_excel_automation()
